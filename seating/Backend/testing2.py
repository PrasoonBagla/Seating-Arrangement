import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Paths to your uploaded Excel files
file_path1 = '13th april.xlsx'
file_path2 = 'courses_to_rooms_with_counts_allocation.xlsx'

# Load the Excel files into pandas DataFrames
df1 = pd.read_excel(file_path1)
df2 = pd.read_excel(file_path2)

# Find the column index for 'Total Course Count' in the first file
# This is to ensure we are highlighting the correct column later on
try:
    total_course_count_col_index = df1.columns.get_loc('Total Course Count') + 1  # openpyxl is 1-indexed
except KeyError:
    raise ValueError("Column 'Total Course Count' not found in the first Excel file.")

# Find differences in 'Total Course Count' column
# This comparison assumes both DataFrames have the same row order and same number of rows
diff_indices = df1[df1['Total Course Count'] != df2['Total Course Count']].index

# Load the first Excel file with openpyxl for highlighting
wb = load_workbook(file_path1)
ws = wb.active

# Highlight color
fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Convert pandas index to Excel row numbers (Excel is 1-indexed and header row is row 1)
excel_rows_to_highlight = [idx + 2 for idx in diff_indices]  # Adjusting for header row

# Highlight the rows with differences in the 'Total Course Count' column
for row in excel_rows_to_highlight:
    ws.cell(row=row, column=total_course_count_col_index).fill = fill

# Save the highlighted Excel file
highlighted_file_path = 'highlighted_13th_april.xlsx'
wb.save(highlighted_file_path)

highlighted_file_path
